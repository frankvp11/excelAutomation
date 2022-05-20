import calendar
from re import I

import pygame, sys
import pandas as pd
from datetime import datetime, timedelta, date
from dateutil import parser
import os
from openpyxl import load_workbook
import json
import shutil


import pygame
import time

pygame.font.init()

text_font = pygame.font.Font("assets/font.ttf", 16)
leftover = 0

class Button():
	def __init__(self, image, pos, text_input, font, base_color, hovering_color):
		self.image = image
		self.x_pos = pos[0]
		self.y_pos = pos[1]
		self.font = font
		self.base_color, self.hovering_color = base_color, hovering_color
		self.text_input = text_input
		self.text = self.font.render(self.text_input, True, self.base_color)
		if self.image is None:
			self.image = self.text
		self.rect = self.image.get_rect(center=(self.x_pos, self.y_pos))
		self.text_rect = self.text.get_rect(center=(self.x_pos, self.y_pos))

	def update(self, screen):
		if self.image is not None:
			screen.blit(self.image, self.rect)
		screen.blit(self.text, self.text_rect)

	def checkForInput(self, position):
		if position[0] in range(self.rect.left, self.rect.right) and position[1] in range(self.rect.top, self.rect.bottom):
			return True
		return False

	def changeColor(self, position):
		if position[0] in range(self.rect.left, self.rect.right) and position[1] in range(self.rect.top, self.rect.bottom):
			self.text = self.font.render(self.text_input, True, self.hovering_color)
		else:
			self.text = self.font.render(self.text_input, True, self.base_color)

class InputBox:

    def __init__(self, x, y, w, h, text=''):
        self.rect = pygame.Rect(x, y, w, h)
        self.color = (255,255,255)
        self.text = text
        self.txt_surface = text_font.render(text, True, self.color)
        self.active = False
        self.score = 1
        # Cursor declare
        self.txt_rect = self.txt_surface.get_rect()
        self.cursor = pygame.Rect(self.txt_rect.topright, (3, self.txt_rect.height + 2))

    def handle_event(self, event, screen):
        if event.type == pygame.MOUSEBUTTONDOWN:
            # If the user clicked on the input_box rect.
            if self.rect.collidepoint(event.pos):
                
                # Toggle the active variable.
                self.active = not self.active
            else:
                self.active = False

        if event.type == pygame.KEYDOWN:
            if self.active:
                if event.key == pygame.K_RETURN:
                    print(self.text)
                    global leftover
                    leftover += self.score
                    self.score = 0
                    self.active = False
                elif event.key == pygame.K_BACKSPACE:
                    try:
                        self.text = self.text[:-1]
                    except TypeError:
                        pass
                else:
                    try:
                        self.text += event.unicode
                    except TypeError:
                        pass
                    # Cursor

                    self.txt_rect.size = self.txt_surface.get_size()
                    self.cursor.topleft = self.txt_rect.topright

                    # Limit characters           -20 for border width
                    if self.txt_surface.get_width() > self.rect.w - 15:
                        self.text = self.text[:-1]

    def draw(self, screen):
        # Blit the text.
        screen.blit(self.txt_surface, (self.rect.x + 5, self.rect.y + 10))
        # Blit the rect.
        pygame.draw.rect(screen, self.color, self.rect, 1)
        # Blit the  cursor
        if time.time() % 1 > 0.5:
            text_rect = self.txt_surface.get_rect(topleft = (self.rect.x + 5, self.rect.y + 10))

            # set cursor position
            self.cursor.midleft = text_rect.midright
            if self.active:
                pygame.draw.rect(screen, self.color, self.cursor)



    def update(self):
        # Re-render the text.
        self.txt_surface = text_font.render(self.text, True, self.color)


def create_new(current_client):
    original = "C:/Users/frank/Desktop/excelAutomation/assets/template.xlsx"
    new = f"C:/Users/frank/Desktop/excelAutomation/clientFiles/{current_client}.xlsx"
    shutil.copyfile(original, new)


def write_json(new_data, filename='data.json'):
    with open(filename,'r+') as file:
          # First we load existing data into a dict.
        file_data = json.load(file)
        # Join new_data with file_data inside emp_details
        file_data["Clients"].append(new_data)
        # Sets file's current position at offset.
        file.seek(0)
        # convert back to json.
        json.dump(file_data, file, indent = 4)
 
 
def write_txt(new_data, filename="clients.txt"):
    with open(filename, 'a') as file:
        file.write(new_data + "\n")


def read_data2(current_client):
    with open("data.json", "r") as file:
        data = json.load(file)
        stack = data["Clients"]
        while stack:
            if current_client in stack[0]:
                return stack[0][current_client]
            stack.pop(0)
            
def read_txt(current_client):
    with open("clients.txt", "r") as file:
        data = file.read()
        if current_client in data:
            return True
    return False

def read_excel(current_client):
    list_of_things_done = []
    list_of_prices = []
    list_of_dates = []
    workbook = load_workbook(f"C:/Users/frank/Desktop/excelAutomation/clientFiles/{current_client}.xlsx", read_only=True)
    worksheet = workbook.active
    for i in range(21, 27):
        list_of_things_done.append(worksheet.cell(row=i, column=3).value)
        list_of_prices.append(worksheet.cell(row=i, column=10).value)
        list_of_dates.append(worksheet.cell(row=i, column=7).value)
    return list_of_things_done, list_of_prices, list_of_dates

def insert_data2(current_price, current_job, current_address, current_client, date_done):
    path = f"C:/Users/frank/Desktop/excelAutomation/clientFiles/{current_client}.xlsx"
    workbook = load_workbook(path)
    clear_file(workbook, path)
    worksheet = workbook.active
    worksheet.cell(10, 2, current_client) 
    worksheet.cell(11, 2, current_address)
    date = datetime.today().strftime('%Y-%m-%d')
    worksheet.cell(10, 10, date)
    i = 21
    while current_price:
        price = current_price.pop(0).text
        try:
            price = int(price)
        except:
            price = price

        job = current_job.pop(0)
        date2 = date_done.pop(0)
        if worksheet.cell(i, 3).value != worksheet.cell(100, 3).value:
            pass
        else:
            worksheet.cell(i, 3, job.text)
            worksheet.cell(i, 7, date2.text)
            worksheet.cell(i, 10, price)
        i += 1 
        workbook.save(path)

def print_file(path):
    os.startfile(path, 'print')
    workbook = load_workbook(path)
    worksheet = workbook.active
    current_value = worksheet.cell(9, 10).value
    list_of_val = list(current_value)
    list_of_val[0] = str(int(list_of_val[0]) + 1)
    new_value = "".join(list_of_val)
    worksheet.cell(9, 10, value=new_value)
    workbook.save(path)

def clear_file(workbook, path):
    worksheet = workbook.active
    for i in range(20, 28):
        worksheet.cell(i, 3).value=None
        worksheet.cell(i, 7).value=None
        worksheet.cell(i, 10).value=None


    
    workbook.save(path)
    return


pygame.init()

SCREEN = pygame.display.set_mode((1280, 720))
pygame.display.set_caption("Menu")

BG = pygame.image.load("assets/Background.png")

def get_font(size): # Returns Press-Start-2P in the desired size
    return pygame.font.Font("assets/font.ttf", size)

def new_client():
    client_name = InputBox(500, 130, 300, 50)
    client_address = InputBox(500, 240, 300, 50)
    client_job_price1 = InputBox(500, 340, 300, 50)
    client_job_description1 = InputBox(500, 440, 300, 50)
    client_job_price2 = InputBox(850, 340, 300, 50)
    client_job_description2 = InputBox(850, 440, 300, 50)
    while True:
        PLAY_MOUSE_POS = pygame.mouse.get_pos()

        SCREEN.fill("black")

        PLAY_TEXT = get_font(20).render("This is the new client screen.", True, "White")
        client_name_text = get_font(20).render("Enter the client's name here", True, "white")
        client_address_text = get_font(20).render("Enter the client's address here", True, 'white')
        name_text_rect = client_name_text.get_rect(center=(640, 100))
        address_text_rect = client_address_text.get_rect(center=(640, 200))
        client_price_text = get_font(20).render("Price:", True, 'white')
        client_price_rect = client_price_text.get_rect(center=(640, 310))
        client_job_desc = get_font(20).render("Job/ Job Description:", True, 'white')
        client_job_desc_rect = client_job_desc.get_rect(center=(640,410))
        PLAY_RECT = PLAY_TEXT.get_rect(center=(640, 50))
        SCREEN.blit(client_job_desc, client_job_desc_rect)
        SCREEN.blit(PLAY_TEXT, PLAY_RECT)
        SCREEN.blit(client_name_text, name_text_rect)
        SCREEN.blit(client_address_text, address_text_rect)
        SCREEN.blit(client_price_text, client_price_rect)

        PLAY_BACK = Button(image=None, pos=(640, 650), 
                            text_input="BACK", font=get_font(50), base_color="White", hovering_color="Green")

        PLAY_BACK.changeColor(PLAY_MOUSE_POS)
        PLAY_BACK.update(SCREEN)
        done_button = Button(image=None, pos=(640, 550), text_input="Save", font=get_font(30), base_color='white', hovering_color='green')
        done_button.changeColor(PLAY_MOUSE_POS)
        done_button.update(SCREEN)
        
 


        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if PLAY_BACK.checkForInput(PLAY_MOUSE_POS):
                    main_menu()
                if done_button.checkForInput(PLAY_MOUSE_POS):
                    current_client = client_name.text
                    current_address = client_address.text
                    current_job_desc = client_job_description1.text
                    current_job_price = client_job_price1.text
                    if client_job_description2.text != "" and client_job_price2.text != "":
                        y = {current_client:{"Address":current_address, 
                                             "Jobs": [current_job_desc, client_job_description2.text],
                                             "Price": [current_job_price, client_job_price2.text]}}
                    else:
                        y = {current_client:{"Address":current_address, 
                                             "Jobs": [current_job_desc],
                                             "Price": [current_job_price]}}
                    write_json(y)
                    write_txt(current_client)
                    create_new(current_client)
                    client_name.text = ""
                    client_address.text = ""
                    client_job_description1.text = ""
                    client_job_price1.text = ""
                    client_job_description2.text = ""
                    client_job_price2.text = ""
            client_name.handle_event(event, SCREEN)
            client_address.handle_event(event, SCREEN)
            client_job_description1.handle_event(event, SCREEN)
            client_job_price1.handle_event(event, SCREEN)
            client_job_description2.handle_event(event, SCREEN)
            client_job_price2.handle_event(event, SCREEN)


        client_job_description1.update()
        client_job_description1.draw(SCREEN)
        client_job_description2.update()
        client_job_description2.draw(SCREEN)
        client_job_price1.update()
        client_job_price1.draw(SCREEN)
        client_job_price2.update()
        client_job_price2.draw(SCREEN)
        client_address.update()
        client_address.draw(SCREEN)
        client_name.update()
        client_name.draw(SCREEN)


        pygame.display.update()
    


def existing_client():
    client_name = InputBox(500, 200, 300, 50)
    is_in_text = False

    while True:
        OPTIONS_MOUSE_POS = pygame.mouse.get_pos()
        SCREEN.fill("black")

        client_name_text = get_font(20).render("Enter the client's name here", True, "white")
        name_text_rect = client_name_text.get_rect(center=(640, 160))
        SCREEN.blit(client_name_text, name_text_rect)
        OPTIONS_TEXT = get_font(15).render("This is the existing clients screen.", True, "white")
        OPTIONS_RECT = OPTIONS_TEXT.get_rect(center=(640, 50))
        SCREEN.blit(OPTIONS_TEXT, OPTIONS_RECT)

        OPTIONS_BACK = Button(image=None, pos=(640, 620), 
                            text_input="BACK", font=get_font(20), base_color="white", hovering_color="Green")
        OPTIONS_BACK.changeColor(OPTIONS_MOUSE_POS)
        OPTIONS_BACK.update(SCREEN)

        save_button = Button(image=None, pos=(640, 350), text_input="Save", font=get_font(20), base_color="white", hovering_color="green")
        save_button.changeColor(OPTIONS_MOUSE_POS)
        save_button.update(SCREEN)


        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if OPTIONS_BACK.checkForInput(OPTIONS_MOUSE_POS):
                    main_menu()
                if save_button.checkForInput(OPTIONS_MOUSE_POS):
                    current_client = client_name.text
                    is_in_text = read_txt(current_client)
                    if is_in_text:
                        change_data_screen(current_client)                   
                    
            client_name.handle_event(event, SCREEN)

        client_name.update()
        client_name.draw(SCREEN)

        pygame.display.update()

def change_data_screen(current_client):
    path = f"C:/Users/frank/Desktop/excelAutomation/clientFiles/{current_client}.xlsx"

    jobs_done, prices_for_jobs, dates = read_excel(current_client)
    dates_dict = []
    jobs_done_dict = []
    prices_for_jobs_dict = []
    
    for i in range(5):
        try:
            dates[i] = parser.parse(dates[i])
            dates_dict.append(InputBox(300, 150+i*50, 250, 50, text=dates[i].strftime("%d %b %Y ")))
        except TypeError:
            dates_dict.append(InputBox(300, 150+i*50, 250, 50))

        try:
            jobs_done_dict.append(InputBox(50, 150+i*50, 250, 50, text=jobs_done[i]))
        except TypeError:
            jobs_done_dict.append(InputBox(50, 150+i*50, 250, 50))

        try:
            prices_for_jobs_dict.append(InputBox(550, 150+i*50, 125, 50, text=str(prices_for_jobs[i])))
        except TypeError:
            prices_for_jobs_dict.append(InputBox(550, 150+i*50, 125, 50))



    jobs_buttons = []
    jobs = read_data2(current_client)['Jobs']
    for i in range(len(jobs)):
        jobs_buttons.append(Button(image=None, pos=(950, 100+i*30), text_input=jobs[i], font=get_font(15), base_color='white', hovering_color='green'))

    cal = calendar.Calendar()
    cal.setfirstweekday(6)
    month = date.today().month
    year = date.today().year
    this_month_data = cal.monthdatescalendar(year, month)
    this_month_data2 = cal.monthdatescalendar(year, month)
    for i in range(len(this_month_data)):
        for j in range(len(this_month_data[i])):
            this_month_data[i][j] = this_month_data[i][j].strftime("%m/%d")
            this_month_data2[i][j] = this_month_data2[i][j].strftime("%d")
    list_of_date_buttons = []
    for i in range(len(this_month_data2)):
        list_of_date_button_j = []
        for j in range(len(this_month_data2[i])):
            list_of_date_button_j.append(Button(image=None, pos=(50+j*100, 500+i*30), text_input=str(this_month_data2[i][j]), font=get_font(25), base_color='white', hovering_color='green'))
        list_of_date_buttons.append(list_of_date_button_j)


    weeks_button_dict = []
    jobs_button_dict = []
    for i in range(5):
        weeks_button_dict.append(Button(image=None, pos=(1150, 175+i*50), text_input=f"{i + 1}", base_color='white', hovering_color='Green', font=get_font(25)))
        jobs_button_dict.append(Button(image=None, pos=(900, 175+i*50), text_input=f"{i + 1}", base_color='white', hovering_color='Green', font=get_font(25)))
    
    
    
    while True:
 


        OPTIONS_MOUSE_POS = pygame.mouse.get_pos()
        SCREEN.fill("black")
        jobs_done_text = get_font(15).render("Job done", True, 'white')
        jobs_done_text_rect = jobs_done_text.get_rect(center=(150, 125))
        jobs_date_text = get_font(15).render("Date done", True, 'white')
        jobs_date_text_rect = jobs_date_text.get_rect(center=(450, 125))
        SCREEN.blit(jobs_date_text, jobs_date_text_rect)
        SCREEN.blit(jobs_done_text, jobs_done_text_rect)
        client_name_text = get_font(20).render(current_client, True, 'white')
        client_name_rect = client_name_text.get_rect(center=(150,50))
        SCREEN.blit(client_name_text, client_name_rect)
        save_button = Button(image=None, pos=(1100, 700), text_input="Save", font=get_font(20), base_color='white', hovering_color='Green')
        save_button.changeColor(OPTIONS_MOUSE_POS)
        save_button.update(SCREEN)
        





        print_button = Button(image=None, pos=(1100, 600), text_input="Print File", font=get_font(20), base_color='white', hovering_color='green')
        print_button.changeColor(OPTIONS_MOUSE_POS)
        print_button.update(SCREEN)
        
        OPTIONS_BACK = Button(image=None, pos=(750, 620), 
                            text_input="BACK", font=get_font(20), base_color="white", hovering_color="Green")
        OPTIONS_BACK.changeColor(OPTIONS_MOUSE_POS)
        OPTIONS_BACK.update(SCREEN)
        change_all_4_wk = Button(image=None, pos=(1150, 100), text_input="4 weeks", base_color='white',
                                hovering_color='Green', font=get_font(15))
                    
        change_all_4_wk.changeColor(OPTIONS_MOUSE_POS)
        change_all_4_wk.update(SCREEN)



        for x in range(len(list_of_date_buttons)):
            for y in range(len(list_of_date_buttons[i])):
                list_of_date_buttons[x][y].changeColor(OPTIONS_MOUSE_POS)
                list_of_date_buttons[x][y].update(SCREEN)


        for i in range(len(weeks_button_dict)):
            weeks_button_dict[i].changeColor(OPTIONS_MOUSE_POS)
            weeks_button_dict[i].update(SCREEN)
            jobs_button_dict[i].changeColor(OPTIONS_MOUSE_POS)
            jobs_button_dict[i].update(SCREEN)
        for i in range(len(dates_dict)):

            dates_dict[i].update()
            dates_dict[i].draw(SCREEN)
            jobs_done_dict[i].update()
            jobs_done_dict[i].draw(SCREEN)
            prices_for_jobs_dict[i].update()
            prices_for_jobs_dict[i].draw(SCREEN)

        for i in range(len(jobs_buttons)):
            jobs_buttons[i].changeColor(OPTIONS_MOUSE_POS)  
            jobs_buttons[i].update(SCREEN)

        


        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if print_button.checkForInput(OPTIONS_MOUSE_POS):
                    insert_data2(prices_for_jobs_dict, jobs_done_dict, read_data2(current_client)["Address"], current_client, dates_dict)
                    print_file(path)
                if OPTIONS_BACK.checkForInput(OPTIONS_MOUSE_POS):
                    existing_client()
                try:
                    if jobs_buttons[0].checkForInput(OPTIONS_MOUSE_POS):
                        jobs_buttons[0].base_color = 'Blue'
                        jobs_buttons[1].base_color = 'White'
                    if jobs_buttons[1].checkForInput(OPTIONS_MOUSE_POS):
                        jobs_buttons[1].base_color = 'Blue'
                        jobs_buttons[0].base_color = 'White'
                except:
                    if jobs_buttons[0].checkForInput(OPTIONS_MOUSE_POS):
                        jobs_buttons[0].base_color = 'Blue'
                
                if save_button.checkForInput(OPTIONS_MOUSE_POS):
                    for i in range(5):
                        try:
                            if prices_for_jobs_dict[i].text[0] == "$":
                                prices_for_jobs_dict[i].text = prices_for_jobs_dict[i].text[1:]
                        except:
                            pass

                    insert_data2(prices_for_jobs_dict, jobs_done_dict, read_data2(current_client)["Address"], current_client, dates_dict)
                    change_data_screen(current_client)
                for i in range(len(weeks_button_dict)):  
                    if weeks_button_dict[i].checkForInput(OPTIONS_MOUSE_POS):  
                        weeks_button_dict[i].base_color = "Blue"
                        for k in range(len(weeks_button_dict)):
                            if k != i and weeks_button_dict[k].base_color == "Blue":
                                weeks_button_dict[k].base_color = 'White'
                        
                if change_all_4_wk.checkForInput(OPTIONS_MOUSE_POS):
                    for i in range(len(weeks_button_dict)):
                        if weeks_button_dict[i].base_color == 'Blue':
                            dates[i] += timedelta(weeks=4)
                            dates_dict[i] = InputBox(300, 150+i*50, 250, 50, text=dates[i].strftime("%d %b %Y"))


                for i in range(len(jobs_button_dict)):
                    if jobs_button_dict[i].checkForInput(OPTIONS_MOUSE_POS):
                        try:
                            if jobs_buttons[0].base_color == 'Blue':
                                ###
                                # changed jobs[0] from jobs_done
                                jobs_done_dict[i] = InputBox(50, 150+i*50, 250, 50, text=jobs[0])
                                prices_for_jobs_dict[i] = InputBox(550, 150+i*50, 125, 50, text=read_data2(current_client)["Price"][0])

                            if jobs_buttons[1].base_color == 'Blue':
                                jobs_done_dict[i] = InputBox(50, 150+i*50, 250, 50, text=jobs[1])
                                prices_for_jobs_dict[i] = InputBox(550, 150+i*50, 125, 50, text=read_data2(current_client)["Price"][1])


                        except IndexError:
                            if jobs_buttons[0].base_color == 'Blue':
                                jobs_done_dict[i] = InputBox(50, 150+i*50, 250, 50, text=jobs[0])
                                prices_for_jobs_dict[i] = InputBox(550, 150+i*50, 125, 50, text=read_data2(current_client)["Price"][0])


                for w in range(len(list_of_date_buttons)):
                    for u in range(len(list_of_date_buttons[w])):
                        if list_of_date_buttons[w][u].checkForInput(OPTIONS_MOUSE_POS):
                            for i in range(len(weeks_button_dict)):
                                if weeks_button_dict[i].base_color == "Blue":
                                    new_u = this_month_data[w][u]
                                    date20 = parser.parse(f"{year}/{new_u}")   
                                    dates[i] = date20
                                    #
                                    dates_dict[i] = InputBox(300, 150+i*50, 250, 50, text=date20.strftime("%d %b %Y"))

            
            for i in range(5):

                dates_dict[i].handle_event(event, SCREEN)
                jobs_done_dict[i].handle_event(event, SCREEN)
                prices_for_jobs_dict[i].handle_event(event, SCREEN)
            

            





        change_dates = get_font(15).render("Change Date", True, 'white')
        change_dates_rect = change_dates.get_rect(center=(1150, 50))
        SCREEN.blit(change_dates, change_dates_rect)
        add_job_to_file = get_font(15).render("Add Jobs", True, 'white')
        add_job_to_file_rect = add_job_to_file.get_rect(center=(900, 50))
        SCREEN.blit(add_job_to_file, add_job_to_file_rect)

            
        pygame.display.update()
    

    

def main_menu():
    while True:
        SCREEN.blit(BG, (0, 0))

        MENU_MOUSE_POS = pygame.mouse.get_pos()

        MENU_TEXT = get_font(100).render("MAIN MENU", True, "#b68f40")
        MENU_RECT = MENU_TEXT.get_rect(center=(640, 100))

        PLAY_BUTTON = Button(image=pygame.image.load("assets/Play Rect.png"), pos=(640, 250), 
                            text_input="New Client", font=get_font(20), base_color="#d7fcd4", hovering_color="White")
        OPTIONS_BUTTON = Button(image=pygame.image.load("assets/Options Rect.png"), pos=(640, 400), 
                            text_input="Existing Clients", font=get_font(20), base_color="#d7fcd4", hovering_color="White")
        QUIT_BUTTON = Button(image=pygame.image.load("assets/Quit Rect.png"), pos=(640, 550), 
                            text_input="QUIT", font=get_font(75), base_color="#d7fcd4", hovering_color="White")

        SCREEN.blit(MENU_TEXT, MENU_RECT)

        for button in [PLAY_BUTTON, OPTIONS_BUTTON, QUIT_BUTTON]:
            button.changeColor(MENU_MOUSE_POS)
            button.update(SCREEN)
        
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                if PLAY_BUTTON.checkForInput(MENU_MOUSE_POS):
                    new_client()
                if OPTIONS_BUTTON.checkForInput(MENU_MOUSE_POS):
                    existing_client()
                if QUIT_BUTTON.checkForInput(MENU_MOUSE_POS):
                    pygame.quit()
                    sys.exit()

        pygame.display.update()

main_menu()