from xml.dom.pulldom import PROCESSING_INSTRUCTION
from openpyxl import Workbook, load_workbook
from os import listdir
from os.path import isfile, join
import time
import json
from helper2 import couldnt_find_client, get_curr_price, get_curr_job
import os
from datetime import datetime


path = "C:/Users/frank/Desktop/excelAutomation/clientFiles"
all_files = [file for file in listdir(path) if isfile(join(path, file))]


def read_excel(current_client):
    list_of_things_done = []
    list_of_prices = []
    list_of_dates = []
    workbook = load_workbook(f"C:/Users/frank/Desktop/excelAutomation/clientFiles/{current_client}.xlsx", read_only=True)
    worksheet = workbook.active
    for i in range(21, 30):
        if worksheet.cell(row=i, column=3).value == None:
            pass
        else:
            list_of_things_done.append(worksheet.cell(row=i, column=3).value)
            list_of_prices.append(worksheet.cell(row=i, column=10).value)
            list_of_dates.append(worksheet.cell(row=i, column=7).value)
    return list_of_things_done, list_of_prices, list_of_dates


def get_curr_client():
    current_client = input("Enter the client's full name whos file you want to change: \n")
    return current_client

def could_find_client(current_client):
    print("Was able to find client")
    current_price = get_curr_price()
    current_job = get_curr_job()
    current_address = read_data2(current_client)["Address"]
    date_job_done = get_date()
    os.system("cls")
    return current_price, current_job, current_address, date_job_done


def get_date():
    date = input("What day did you do this on? \n")
    return date

def insert_data(current_price, current_job, current_address, current_client, date_done):
    path = f"C:/Users/frank/Desktop/excelAutomation/clientFiles/{current_client}.xlsx"
    workbook = load_workbook(path)
    clear_file(workbook, path)

    worksheet = workbook.active
    worksheet.cell(10, 2, current_client) 
    worksheet.cell(11, 2, current_address)
    date = datetime.today().strftime('%Y-%m-%d')
    worksheet.cell(10, 10, date)
    print(worksheet.cell(21, 3).value)
    if worksheet.cell(21, 3).value == None:
        worksheet.cell(21, 3, current_job)
        worksheet.cell(21, 10, current_price)
        worksheet.cell(21, 7, date_done)
    else:
        if worksheet.cell(22, 3).value == None or worksheet.cell(22, 3).value == "":
            worksheet.cell(22, 3, current_job)
            worksheet.cell(22, 10, current_price)
            worksheet.cell(22, 7, date_done)

        else:
            if worksheet.cell(23, 3).value == None or worksheet.cell(22, 3).value == "":
                worksheet.cell(23, 3, current_job)
                worksheet.cell(23, 10, current_price)
                worksheet.cell(23, 7, date_done)

            else:
                if worksheet.cell(24, 3).value == None:                
                    worksheet.cell(24, 3, current_job)
                    worksheet.cell(24, 10, current_price)
                    worksheet.cell(24, 7, date_done)

                    
    print("Done inserting data")
    workbook.save(path)  


def insert_data2(current_price, current_job, current_address, current_client, date_done):
    path = f"C:/Users/frank/Desktop/excelAutomation/clientFiles/{current_client}.xlsx"
    workbook = load_workbook(path)
    clear_file(workbook, path)
    worksheet = workbook.active
    worksheet.cell(10, 2, current_client) 
    worksheet.cell(11, 2, current_address)
    date = datetime.today().strftime('%Y-%m-%d')
    worksheet.cell(10, 10, date)
    i = 21
    while current_price:
        price = current_price.pop(0).text
        try:
            price = int(price)
        except:
            price = price
            # print("Not possible")
            # print(price)
        job = current_job.pop(0)
        date2 = date_done.pop(0)
        if worksheet.cell(i, 3).value != worksheet.cell(100, 3).value:
            pass
        else:
            worksheet.cell(i, 3, job.text)
            worksheet.cell(i, 7, date2.text)
            worksheet.cell(i, 10, price)
        i += 1 
        workbook.save(path)

    



def clear_file(workbook, path):
    worksheet = workbook.active
    for i in range(20, 28):
        worksheet.cell(i, 3).value=None
        worksheet.cell(i, 7).value=None
        worksheet.cell(i, 10).value=None


    
    workbook.save(path)
    print("Cleared file")
    return

                


def read_data2(current_client):
    with open("data.json", "r") as file:
        data = json.load(file)
        stack = data["Clients"]
        while stack:
            if current_client in stack[0]:
                return stack[0][current_client]
            stack.pop(0)
            

def read_txt(current_client):
    with open("clients.txt", "r") as file:
        data = file.read()
        if current_client in data:
            return True
    return False




def interrogate():
    current_client = get_curr_client()
    boolean_client_in_text = read_txt(current_client)
    if f"{current_client}.xlsx" in all_files or boolean_client_in_text:
        current_price, current_job, current_address, date_done = could_find_client(current_client)
  
        insert_data(current_price, current_job, current_address, current_client, date_done)
    else:
        couldnt_find_client(current_client)
    interrogate()

#interrogate()
def print_file(path):
    os.startfile(path, 'print')